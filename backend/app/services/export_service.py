import io
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.models.affectation import Affectation, StatutAffectation
from app.models.lieu import Lieu


def _header_fill():
    return PatternFill(start_color="0B3D3D", end_color="0B3D3D", fill_type="solid")


def _logo_path() -> str | None:
    if settings.logo_path and Path(settings.logo_path).exists():
        return settings.logo_path
    for candidate in (
        Path(__file__).resolve().parent.parent / "assets" / "logo-region-ouest.png",
        Path(__file__).resolve().parent.parent.parent.parent / "frontend" / "public" / "logo-region-ouest.png",
    ):
        if candidate.exists():
            return str(candidate)
    return None


def _pdf_branding(elements, styles, subtitle: str = ""):
    logo = _logo_path()
    if logo:
        elements.append(Image(logo, width=2.2 * cm, height=2.2 * cm))
        elements.append(Spacer(1, 0.2 * cm))
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], textColor=colors.HexColor("#0B3D3D"))
    elements.append(Paragraph("CONSEIL REGIONAL DE L'OUEST", title_style))
    elements.append(Paragraph("Region de l'Ouest — Republique du Cameroun", styles["Normal"]))
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Heading2"]))
    elements.append(Paragraph(f"Document genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 0.4 * cm))


def export_materiels_excel(materiels: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaire Materiel"
    headers = ["Matricule", "Designation", "Categorie", "Marque", "Modele", "N Serie", "Etat", "Quantite", "Seuil alerte", "Valeur FCFA"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
    for m in materiels:
        ws.append([
            m.matricule, m.designation, m.categorie.value, m.marque or "", m.modele or "",
            m.numero_serie or "", m.etat.value, m.quantite, m.seuil_alerte or "", m.valeur_acquisition or "",
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_affectations_excel(affectations: list[Affectation]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Affectations"
    headers = ["Matricule", "Materiel", "Lieu", "Beneficiaire", "Raison", "Statut", "Date debut", "Date fin"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
    for a in affectations:
        ws.append([
            a.materiel.matricule if a.materiel else "",
            a.materiel.designation if a.materiel else "",
            a.lieu.nom if a.lieu else "",
            a.beneficiaire, a.raison, a.statut.value,
            a.date_debut.strftime("%d/%m/%Y") if a.date_debut else "",
            a.date_fin.strftime("%d/%m/%Y") if a.date_fin else "",
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_materiels_pdf(materiels: list, titre: str = "Inventaire du patrimoine") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1.5 * cm, leftMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = []
    _pdf_branding(elements, styles, titre)
    data = [["Matricule", "Designation", "Categorie", "Etat", "Stock", "Seuil"]]
    for m in materiels:
        data.append([
            m.matricule, m.designation[:36], m.categorie.value, m.etat.value,
            str(m.quantite), str(m.seuil_alerte or "—"),
        ])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D3D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F5F0")]),
    ]))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


def export_bon_affectation_pdf(affectation: Affectation) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2 * cm, leftMargin=2 * cm, topMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    m = affectation.materiel
    l = affectation.lieu
    elements = []
    _pdf_branding(elements, styles, "BON D'AFFECTATION DE MATERIEL")
    elements.extend([
        Paragraph(f"<b>Reference :</b> AFF-{affectation.id:05d}", styles["Normal"]),
        Paragraph(f"<b>Date d'affectation :</b> {affectation.date_debut.strftime('%d/%m/%Y')}", styles["Normal"]),
        Spacer(1, 0.4 * cm),
        Paragraph("<b>MATERIEL AFFECTE</b>", styles["Heading3"]),
        Paragraph(f"Designation : {m.designation if m else '-'}", styles["Normal"]),
        Paragraph(f"Matricule : {m.matricule if m else '-'} | N° serie : {m.numero_serie or '-'}", styles["Normal"]),
        Paragraph(f"Categorie : {m.categorie.value if m else '-'} | Etat : {m.etat.value if m else '-'}", styles["Normal"]),
        Spacer(1, 0.3 * cm),
        Paragraph("<b>DESTINATION</b>", styles["Heading3"]),
        Paragraph(f"Structure : {l.nom if l else '-'} ({l.type_lieu.value if l else '-'})", styles["Normal"]),
        Paragraph(f"Beneficiaire / Responsable : {affectation.beneficiaire}", styles["Normal"]),
        Spacer(1, 0.3 * cm),
        Paragraph(f"<b>Motif :</b> {affectation.raison}", styles["Normal"]),
    ])
    if affectation.document_reference:
        elements.append(Paragraph(f"<b>Reference document :</b> {affectation.document_reference}", styles["Normal"]))
    if affectation.signataire_nom:
        elements.append(Paragraph(
            f"<b>Signe par :</b> {affectation.signataire_nom} le "
            f"{affectation.date_signature.strftime('%d/%m/%Y') if affectation.date_signature else ''}",
            styles["Normal"],
        ))
    elements.append(Spacer(1, 1.2 * cm))
    sig_table = Table(
        [
            ["Le responsable CRO", "Le beneficiaire"],
            ["", ""],
            ["Date : _______________", "Date : _______________"],
        ],
        colWidths=[8 * cm, 8 * cm],
        rowHeights=[0.6 * cm, 2 * cm, 0.6 * cm],
    )
    sig_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(
        "<i>Ce document atteste de la mise a disposition du materiel cite ci-dessus au beneficiaire designe.</i>",
        styles["Normal"],
    ))
    doc.build(elements)
    return buffer.getvalue()


def rapport_par_lieu_pdf(db: Session) -> bytes:
    lieux = db.query(Lieu).all()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    _pdf_branding(elements, styles, "Rapport par structure")
    for lieu in lieux:
        affs = (
            db.query(Affectation)
            .options(joinedload(Affectation.materiel))
            .filter(Affectation.lieu_id == lieu.id, Affectation.statut == StatutAffectation.ACTIVE)
            .all()
        )
        if not affs:
            continue
        elements.append(Paragraph(f"<b>{lieu.nom}</b> ({lieu.type_lieu.value}) — {len(affs)} materiel(s)", styles["Heading3"]))
        for a in affs:
            elements.append(Paragraph(
                f"  • {a.materiel.designation if a.materiel else '?'} "
                f"[{a.materiel.matricule if a.materiel else '?'}] → {a.beneficiaire}",
                styles["Normal"],
            ))
        elements.append(Spacer(1, 0.3 * cm))
    doc.build(elements)
    return buffer.getvalue()
