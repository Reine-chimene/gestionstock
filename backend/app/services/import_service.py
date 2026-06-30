import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.constants.catalogues import CATEGORIES_MATERIEL, TYPES_LIEU
from app.models.historique import ActionHistorique, TypeEntite
from app.models.lieu import Lieu, TypeLieu
from app.models.materiel import CategorieMateriel, EtatMateriel, Materiel
from app.services.storage_service import log_historique


def materiel_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Materiels"
    ws.append([
        "matricule", "designation", "categorie", "marque", "modele",
        "numero_serie", "etat", "quantite", "seuil_alerte", "valeur_acquisition",
    ])
    ws.append([
        "CRO-001", "Ordinateur portable", "informatique", "HP", "ProBook",
        "SN123", "neuf", "1", "2", "450000",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def lieux_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lieux"
    ws.append(["nom", "type_lieu", "ville", "adresse", "responsable", "telephone", "email"])
    ws.append(["Lycee de Bafoussam", "lycee", "Bafoussam", "Quartier Tamdja", "Directeur", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row: dict, *keys, default=None):
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return row[k]
    return default


def import_materiels_excel(db: Session, content: bytes, user_id: int) -> dict:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"created": 0, "skipped": 0, "errors": ["Fichier vide ou sans donnees."]}

    headers = [str(h or "").strip().lower().replace(" ", "_") for h in rows[0]]
    created = skipped = 0
    errors: list[str] = []

    for i, values in enumerate(rows[1:], start=2):
        if not any(values):
            continue
        row = dict(zip(headers, values))
        matricule = str(_cell(row, "matricule") or "").strip()
        designation = str(_cell(row, "designation") or "").strip()
        if not matricule or not designation:
            errors.append(f"Ligne {i}: matricule et designation obligatoires.")
            continue

        if db.query(Materiel).filter(Materiel.matricule == matricule).first():
            skipped += 1
            continue

        categorie = str(_cell(row, "categorie", default="autre") or "autre").strip().lower()
        if categorie not in CATEGORIES_MATERIEL:
            categorie = "autre"

        etat = str(_cell(row, "etat", default="neuf") or "neuf").strip().lower()
        try:
            etat_enum = EtatMateriel(etat)
        except ValueError:
            etat_enum = EtatMateriel.NEUF

        try:
            quantite = int(_cell(row, "quantite", default=1) or 1)
        except (TypeError, ValueError):
            quantite = 1

        seuil = _cell(row, "seuil_alerte")
        seuil_alerte = int(seuil) if seuil not in (None, "") else None

        valeur = _cell(row, "valeur_acquisition", "valeur")
        try:
            valeur_acquisition = float(valeur) if valeur not in (None, "") else None
        except (TypeError, ValueError):
            valeur_acquisition = None

        materiel = Materiel(
            matricule=matricule,
            designation=designation,
            categorie=CategorieMateriel(categorie),
            marque=str(_cell(row, "marque") or "") or None,
            modele=str(_cell(row, "modele") or "") or None,
            numero_serie=str(_cell(row, "numero_serie", "n_serie") or "") or None,
            etat=etat_enum,
            quantite=max(1, quantite),
            seuil_alerte=seuil_alerte,
            valeur_acquisition=valeur_acquisition,
        )
        db.add(materiel)
        db.flush()
        log_historique(
            db, TypeEntite.MATERIEL, materiel.id, ActionHistorique.CREATION,
            f"Import Excel : {matricule} — {designation}", user_id,
        )
        created += 1

    db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "errors": errors[:20],
        "message": f"{created} materiel(s) importe(s), {skipped} ignore(s) (doublon).",
    }


def import_lieux_excel(db: Session, content: bytes, user_id: int) -> dict:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"created": 0, "skipped": 0, "errors": ["Fichier vide ou sans donnees."]}

    headers = [str(h or "").strip().lower().replace(" ", "_") for h in rows[0]]
    created = skipped = 0
    errors: list[str] = []

    for i, values in enumerate(rows[1:], start=2):
        if not any(values):
            continue
        row = dict(zip(headers, values))
        nom = str(_cell(row, "nom") or "").strip()
        if not nom:
            errors.append(f"Ligne {i}: nom obligatoire.")
            continue

        if db.query(Lieu).filter(Lieu.nom.ilike(nom)).first():
            skipped += 1
            continue

        type_lieu = str(_cell(row, "type_lieu", default="autre") or "autre").strip().lower()
        if type_lieu not in TYPES_LIEU:
            type_lieu = "autre"

        lieu = Lieu(
            nom=nom,
            type_lieu=TypeLieu(type_lieu),
            ville=str(_cell(row, "ville") or "") or None,
            adresse=str(_cell(row, "adresse") or "") or None,
            responsable=str(_cell(row, "responsable") or "") or None,
            telephone=str(_cell(row, "telephone", "tel") or "") or None,
            email=str(_cell(row, "email") or "") or None,
        )
        db.add(lieu)
        db.flush()
        log_historique(
            db, TypeEntite.LIEU, lieu.id, ActionHistorique.CREATION,
            f"Import Excel : {nom}", user_id,
        )
        created += 1

    db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "errors": errors[:20],
        "message": f"{created} lieu(x) importe(s), {skipped} ignore(s) (doublon).",
    }
